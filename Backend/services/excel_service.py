from openpyxl import load_workbook

EXCEL_FILE = r"C:\Users\Utkarsh\Downloads\app\Intern Attendance Sheet Format __ Utkarsh.xlsx"


def update_excel(attendance):

    wb = load_workbook(EXCEL_FILE)

    ws = wb["June"]

    for row in range(2, ws.max_row + 1):

        excel_date = ws[f"A{row}"].value

        if excel_date is None:
            continue

        if hasattr(excel_date, "date"):
            excel_date = excel_date.date()

        if excel_date == attendance.date:

            if attendance.check_in:
                ws[f"C{row}"] = attendance.check_in.strftime("%H:%M")

            if attendance.check_out:
                ws[f"D{row}"] = attendance.check_out.strftime("%H:%M")

            if attendance.approved:
                ws[f"G{row}"] = "Yes"

            break

    wb.save(EXCEL_FILE)